import mysql.connector
import requests as req
import json
from openpyxl import Workbook,load_workbook,chart
import configparser
import configParserMaker
class SQL():
  def __init__(self, host="localhost",user="root",password='',database="cards"):
    configParserMaker.maker(host, user, password, database)
    config= configparser.ConfigParser()
    config.read('config.ini')
    self.host = config['default']['host']
    self.user = config['default']['username']
    self.password = config['default']['password']
    self.db = config['default']['database']
    self.mydb = mysql.connector.connect(host=self.host,user=self.user,password=self.password,database=self.db)
   
  def searchFun(self,Key):
    mycursor =self.mydb.cursor()
    mycursor.execute("SELECT `name` FROM `cards` WHERE `name` LIKE \""+Key+"%\"")
    d=[]
    for x in mycursor:
      d.append(x[0])
    return d
  def refSerch(self,IDK):
    if self.searchFun(IDK)==[]:
      Keid= IDK.replace("s","'s", IDK.find("s"))
      test=self.searchFun(Keid)
      test.append(self.searchFun(IDK))
      return test
    else:
      return self.searchFun(IDK)
  def sendCard(self,name):
        print(name)
        mycursor=self.mydb.cursor()
        mycursor.execute('INSERT INTO `scanedcards` (`name`) Values ("{}")'.format(name))
        self.mydb.commit()

class Xcel:
  def __init__(self):
    try:
      self.wb = load_workbook("sample.xlsx")
    except:
      self.wb = Workbook()
    # grab the active worksheet
    self.ws = self.wb.active
  # Data can be assigned directly to cells
  
  def InPutData(self,ID,name):
    self.ws.append([ID,name])
  # Save the file
    self.wb.save("sample.xlsx")

    
class Scryfall():
  def __init__(self,name):
    self.name = name
    self.parM = {'q':str(self.name)}
    self.reqa = req.api.get("https://api.scryfall.com/cards/search",params=self.parM)
    if self.reqa.status_code!= 200:
      raise ConnectionAbortedError('Cannot fetch all tasks: {}'.format(self.reqa.status_code))
    self.realName = self.reqa.json()['data'][0]['name']
    self.count = self.reqa.json()['total_cards']
